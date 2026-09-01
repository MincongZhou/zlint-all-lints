#!/usr/bin/env python3
"""
run_all.py —— 对单个证书依次跑完项目里的 4 个分析脚本，并把全部结果写入一个 xlsx 大表

执行的脚本与结果去向:
    1. run_zlint.py   (check_certs_python/)         zlint 全部 433 条规则 → 读取其输出的 CSV 写入 sheet「zlint」
    2. extract_org.py (extract_CertInfo_python/)    组织名（type=组织名）  → sheet「组织名」
    3. extract_sct.py (extract_CertInfo_python/)    SCT 时间（type=SCT时间）→ sheet「SCT时间」
    4. check_ocsp.py  (check_certs_python/)         OCSP 查询（type=OCSP查询）→ sheet「OCSP查询」

xlsx 共 5 个 sheet: 上述 4 个分表 + 1 个「汇总」大表（四表合一，统一三列: type / 内容 / status）。

依赖: python3 + cryptography + openpyxl；zlint-all-lints 需已编译（go build -o zlint-all-lints .）。
      openpyxl 未安装时请先安装:
          python3 -m pip install --user openpyxl      # 或  sudo apt install python3-openpyxl

用法:
    python3 run_all.py <证书路径|证书目录> [输出目录]
    python3 run_all.py                                        # 无参数 → 交互式输入
    证书路径: 单个 *.pem / *.crt / *.cer / *.der（PEM/DER 自动识别，扩展名不限）
    证书目录: 传目录则批量跑其中所有证书（递归查找 .pem/.crt/.cer/.der）
    输出目录: 单个证书默认 results/<证书名>/；批量默认 results/（每个证书再建子目录）
              生成 <证书名>_report.xlsx 及 lint 的 JSON/CSV/JSONL

批量输出结构示例:
    results/
    ├── baidu/                 # 每个证书一个子目录
    │   ├── baidu_report.xlsx  # 5 sheet 大表
    │   └── baidu.json/.csv/.jsonl
    └── example_com/
        └── example_com_report.xlsx
"""

import csv
import glob
import os
import re
import shutil
import subprocess
import sys
import tempfile

# --------------------------------------------------------------------------
# 路径与常量
# --------------------------------------------------------------------------
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
RUN_ZLINT = os.path.join(PROJECT_ROOT, "check_certs_python", "run_zlint.py")
CHECK_OCSP = os.path.join(PROJECT_ROOT, "check_certs_python", "check_ocsp.py")
EXTRACT_ORG = os.path.join(PROJECT_ROOT, "extract_CertInfo_python", "extract_org.py")
EXTRACT_SCT = os.path.join(PROJECT_ROOT, "extract_CertInfo_python", "extract_sct.py")

SHEET_ZLINT = "zlint"
SHEET_ORG = "组织名"
SHEET_SCT = "SCT时间"
SHEET_OCSP = "OCSP查询"


def err(msg):
    print(f"错误: {msg}", file=sys.stderr)


# --------------------------------------------------------------------------
# 步骤 1: run_zlint.py —— 用其输出的 CSV 作为 sheet「zlint」数据
# --------------------------------------------------------------------------
def run_zlint(tmp_dir, out_dir, stem):
    """调用 run_zlint.py（目录级），跑完读取 <stem>.csv，返回 (rows, exit_code)"""
    cmd = [sys.executable, RUN_ZLINT, tmp_dir, out_dir, "--jsonl"]
    print(f"执行: {' '.join(cmd)}")
    rc = subprocess.call(cmd)   # 不捕获，进度实时显示

    csv_path = os.path.join(out_dir, f"{stem}.csv")
    rows = []
    if rc == 0 and os.path.isfile(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header:
                rows = [header] + list(reader)
        print(f"读取 lint CSV: {csv_path}（表头+{len(rows)-1} 行规则）")
    else:
        err(f"run_zlint.py 失败 (exit {rc})，或未找到 {csv_path}")
    return rows, rc


# --------------------------------------------------------------------------
# 步骤 2: extract_org.py —— 解析 stdout 得到 subject 与 O 字段
# --------------------------------------------------------------------------
def run_extract_org(cert_path, der_args):
    proc = subprocess.run(
        [sys.executable, EXTRACT_ORG, cert_path] + der_args,
        capture_output=True, text=True)
    print(proc.stdout, end="")
    if proc.returncode != 0:
        err(f"extract_org.py 失败: {proc.stderr.strip()}")
        return [], proc.returncode

    subject = orgs = ""
    m = re.search(r"^subject:\s*(.+)$", proc.stdout, re.M)
    if m:
        subject = m.group(1).strip()
    m = re.search(r"^O 字段:\s*(.+)$", proc.stdout, re.M)
    if m:
        orgs = m.group(1).strip()
    rows = [["type", "subject", "组织名"], [SHEET_ORG, subject, orgs]]
    return rows, 0


# --------------------------------------------------------------------------
# 步骤 3: extract_sct.py —— 解析 stdout 得到每条 SCT 的时间戳等信息
# --------------------------------------------------------------------------
SCT_RE = re.compile(
    r"\[\d+\]\s*Timestamp:\s*(.+?)\s*\n"
    r"\s*epoch ms:\s*(\d+)\s*\n"
    r"\s*Log ID:\s*([0-9A-F]+)\s*\n"
    r"\s*version:\s*(\S+)",
    re.M)


def run_extract_sct(cert_path, der_args):
    proc = subprocess.run(
        [sys.executable, EXTRACT_SCT, cert_path] + der_args,
        capture_output=True, text=True)
    print(proc.stdout, end="")
    if proc.returncode != 0:
        err(f"extract_sct.py 失败: {proc.stderr.strip()}")
        return [], proc.returncode

    rows = [["type", "序号", "timestamp_utc", "epoch_ms", "log_id", "version"]]
    for i, m in enumerate(SCT_RE.finditer(proc.stdout), 1):
        rows.append([SHEET_SCT, i, m.group(1), int(m.group(2)),
                     m.group(3), m.group(4)])
    if len(rows) == 1:
        rows.append([SHEET_SCT, "无", "该证书没有 CT Precertificate SCTs 扩展", "", "", ""])
    return rows, 0


# --------------------------------------------------------------------------
# 步骤 4: check_ocsp.py —— 只取状态（--status），失败不中断
# --------------------------------------------------------------------------
def run_check_ocsp(cert_path, der_args):
    proc = subprocess.run(
        [sys.executable, CHECK_OCSP, cert_path] + der_args + ["--status"],
        capture_output=True, text=True)
    if proc.returncode == 0:
        status = (proc.stdout.strip().splitlines() or ["?"])[0]
        print(f"OCSP 状态: {status}")
        return [["type", "状态", "详情"], [SHEET_OCSP, status, ""]], 0
    detail = (proc.stderr.strip().splitlines() or [""])[0]
    print(f"OCSP 查询失败: {detail}")
    return [["type", "状态", "详情"], [SHEET_OCSP, "查询失败", detail]], proc.returncode


# --------------------------------------------------------------------------
# 汇总大表：四个 sheet 统一为三列 (type, 内容, status)
#    zlint   -> type=规则类别, 内容=规则名(name),    status=规则状态
#    组织名  -> type=组织名,   内容=空,              status=组织名
#    SCT时间 -> type=SCT时间,  内容=timestamp_utc,   status=空
#    OCSP查询-> type=OCSP查询, 内容=空,              status=状态
# --------------------------------------------------------------------------
_MASTER_STATUS_COL = {
    SHEET_ZLINT: "status",
    SHEET_ORG: "组织名",
    SHEET_SCT: "timestamp_utc",
    SHEET_OCSP: "状态",
}


def build_master(rows_list):
    """rows_list: list[(source, rows)]；rows 首行为表头，返回统一三列的汇总大表。"""
    master = [["type", "内容", "status"]]
    for name, rows in rows_list:
        if len(rows) < 2 or name not in _MASTER_STATUS_COL:
            continue
        hs = rows[0]

        def idx(col):
            return hs.index(col) if col in hs else None

        c_type = idx("type")
        # 内容列: zlint 填规则名(name)，SCT时间 填 timestamp_utc，其余来源留空
        c_content = idx("name") if name == SHEET_ZLINT \
            else (idx("timestamp_utc") if name == SHEET_SCT else None)
        # status 列: SCT时间的值已放内容列，故留空；其余来源填对应值
        c_status = None if name == SHEET_SCT else idx(_MASTER_STATUS_COL[name])

        for row in rows[1:]:
            def get(i):
                return row[i] if i is not None and i < len(row) else ""
            master.append([get(c_type), get(c_content), get(c_status)])
    return master


# --------------------------------------------------------------------------
# xlsx 大表生成（openpyxl）
# --------------------------------------------------------------------------
def write_xlsx(path, sheets):
    """sheets: list[(sheet_name, rows)]，rows 首行为表头（加粗），写入多工作表 xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)   # 去掉默认空 sheet

    header_font = Font(bold=True)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name)
        for r, row in enumerate(rows, 1):
            for c, v in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if r == 1:
                    cell.font = header_font
        # 自动列宽（按内容长度，最长 60）
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    wb.save(path)


# --------------------------------------------------------------------------
# 单证书全流程
# --------------------------------------------------------------------------
def run_one(cert_path, out_dir=None):
    """跑单个证书：4 个脚本 + 生成 xlsx 大表。返回 (ok, xlsx_path)。"""
    stem = os.path.splitext(os.path.basename(cert_path))[0]
    out_dir = out_dir or os.path.join(PROJECT_ROOT, "results", stem)
    os.makedirs(out_dir, exist_ok=True)

    # PEM/DER 自动识别
    der_args = []
    with open(cert_path, "rb") as f:
        head = f.read(256)
    if head.lstrip().startswith(b"-----BEGIN"):
        print(f"输入: {cert_path} (PEM)")
    else:
        print(f"输入: {cert_path} (非 PEM，按 DER 处理)")
        der_args = ["--der"]

    # 临时目录：run_zlint.py 以目录为输入，复制证书（统一 .pem 后缀）进去单独跑
    tmp_dir = tempfile.mkdtemp(prefix=".run_all_", dir=os.path.join(PROJECT_ROOT, "results"))
    try:
        shutil.copy(cert_path, os.path.join(tmp_dir, f"{stem}.pem"))

        results = []   # (sheet_name, rows, exit_code)
        ok = True

        print("\n===== [1/4] run_zlint.py（zlint 全部规则 → JSON/CSV/JSONL） =====")
        rows, rc = run_zlint(tmp_dir, out_dir, stem)
        results.append((SHEET_ZLINT, rows, rc))
        ok &= rc == 0 and len(rows) > 1

        print("\n===== [2/4] extract_org.py（组织名 O 字段） =====")
        rows, rc = run_extract_org(cert_path, der_args)
        results.append((SHEET_ORG, rows, rc))
        ok &= rc == 0

        print("\n===== [3/4] extract_sct.py（SCT 时间） =====")
        rows, rc = run_extract_sct(cert_path, der_args)
        results.append((SHEET_SCT, rows, rc))
        ok &= rc == 0

        print("\n===== [4/4] check_ocsp.py（OCSP 状态，联网） =====")
        rows, rc = run_check_ocsp(cert_path, der_args)
        results.append((SHEET_OCSP, rows, rc))
        ok &= rc == 0

        # 生成 xlsx 大表：4 个分表 + 汇总大表（四表合一）
        xlsx_path = os.path.join(out_dir, f"{stem}_report.xlsx")
        sheet_entries = [(name, r) for name, r, _ in results]
        sheet_entries.append(("汇总", build_master(sheet_entries)))
        write_xlsx(xlsx_path, sheet_entries)
        print(f"\n已生成 xlsx 大表: {xlsx_path}")

        # 汇总
        print("\n================ 汇总 ================")
        print(f"证书: {cert_path}")
        print(f"输出: {out_dir}/")
        print("----------------------------------------")
        for name, rows, rc in results:
            data_rows = len(rows) - 1 if rows else 0
            tag = "[OK]  " if rc == 0 and data_rows > 0 else "[失败] "
            print(f"  {tag}{name}: sheet {data_rows} 行数据 (exit={rc})")
        print("----------------------------------------")
        print(f"xlsx 大表: {xlsx_path}")
        return ok, xlsx_path
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# --------------------------------------------------------------------------
# 主流程：单个证书 or 批量目录
# --------------------------------------------------------------------------
CERT_EXTS = (".pem", ".crt", ".cer", ".der", ".cert")


def run_target(target, out_root=None):
    """依赖检查 + 单个/批量分发（target 已展开 ~）"""
    if not os.path.exists(target):
        err(f"路径不存在 -> {target}")
        sys.exit(1)
    for s in (RUN_ZLINT, CHECK_OCSP, EXTRACT_ORG, EXTRACT_SCT):
        if not os.path.isfile(s):
            err(f"找不到 {s}")
            sys.exit(1)

    # 依赖检查：openpyxl
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        err("缺少 openpyxl，请先安装:\n"
            "    python3 -m pip install --user openpyxl\n"
            "或: sudo apt install python3-openpyxl")
        sys.exit(1)

    if os.path.isdir(target):
        # ---------- 批量：遍历目录下所有证书 ----------
        certs = sorted(
            p for p in glob.glob(os.path.join(target, "**", "*"), recursive=True)
            if os.path.isfile(p) and p.lower().endswith(CERT_EXTS))
        if not certs:
            err(f"目录下没有找到证书文件（{CERT_EXTS}）-> {target}")
            sys.exit(1)
        print(f"批量模式: 发现 {len(certs)} 个证书")
        ok_all, ok, fail = True, 0, []
        for i, c in enumerate(certs, 1):
            stem = os.path.splitext(os.path.basename(c))[0]
            out_dir = out_root or os.path.join(PROJECT_ROOT, "results")
            print(f"\n{'='*60}\n[{i}/{len(certs)}] {c}\n{'='*60}")
            ok_one, _ = run_one(c, os.path.join(out_dir, stem))
            ok_all &= ok_one
            if ok_one:
                ok += 1
            else:
                fail.append(c)
        print(f"\n============ 批量完成 ============")
        print(f"成功 {ok}/{len(certs)}，失败 {len(fail)} 个")
        for c in fail:
            print(f"  [失败] {c}")
        sys.exit(0 if ok_all else 1)
    else:
        # ---------- 单个证书 ----------
        ok, _ = run_one(target, out_root)
        sys.exit(0 if ok else 1)


def interactive():
    """无参数时的交互式输入：证书路径/目录必填，输出目录可回车跳过（参考 run_zlint.py）"""
    print("=== 交互模式（直接回车使用默认值，输入 q 退出）===")

    # 证书路径或目录：必填，循环直到存在（支持 ~ 展开）
    target = os.path.expanduser(input("证书路径或目录: ").strip())
    while True:
        if target.lower() in ("q", "quit"):
            sys.exit(0)
        if os.path.exists(target):
            break
        print(f"  !! 路径不存在: {target}")
        target = os.path.expanduser(input("请重新输入证书路径或目录 (q 退出): ").strip())

    # 输出目录：可选，回车跳过 → 用默认值
    out_dir = os.path.expanduser(input("输出目录 (回车用默认): ").strip())
    if out_dir.lower() in ("q", "quit"):
        sys.exit(0)

    run_target(target, out_dir or None)


def main():
    args = sys.argv[1:]
    if not args:                    # 没有任何参数 → 交互模式
        interactive()
        return
    if args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    target = os.path.expanduser(args[0])
    out_root = os.path.expanduser(args[1]) if len(args) > 1 else None
    run_target(target, out_root)


if __name__ == "__main__":
    main()
